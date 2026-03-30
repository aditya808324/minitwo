# reports/excel.py — Excel report generator for Salon Bot v3

import logging
import os
from datetime import date, datetime

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("[REPORTS] openpyxl not installed — run: pip install openpyxl")


def generate_report(output_path: str = None, date_from: str = None, date_to: str = None) -> str:
    """
    Generate Excel report of bookings.
    Returns the file path of the generated report.
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("[REPORTS] openpyxl not available")
        return None

    from database.schema import get_db, get_setting

    os.makedirs("reports", exist_ok=True)

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"reports/salon_report_{ts}.xlsx"

    salon = get_setting("salon_name", "Salon")
    cur   = get_setting("currency", "₹")

    # ── Fetch data ────────────────────────────────────────────────────
    with get_db() as db:
        query = "SELECT * FROM bookings WHERE status != 'cancelled'"
        params = []
        if date_from:
            query  += " AND date >= ?"
            params.append(date_from)
        if date_to:
            query  += " AND date <= ?"
            params.append(date_to)
        query += " ORDER BY date, slot"
        bookings = db.execute(query, params).fetchall()

        summary = db.execute("""
            SELECT
              COUNT(*) as total,
              COALESCE(SUM(total_price), 0) as gross,
              COALESCE(SUM(advance_amount), 0) as advance_collected,
              SUM(CASE WHEN payment_status='paid' THEN 1 ELSE 0 END) as paid_count,
              SUM(CASE WHEN conflict_flag=1 THEN 1 ELSE 0 END) as conflicts
            FROM bookings WHERE status != 'cancelled'
        """).fetchone()

    # ── Create workbook ───────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Colors
    GOLD   = "D4A017"
    DARK   = "1A0A00"
    LIGHT  = "FBF5E6"
    GREEN  = "2E7D52"
    RED    = "B22222"
    WHITE  = "FFFFFF"
    GRAY   = "F5F2EC"

    def _fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _border():
        thin = Side(style="thin", color="E0D8C8")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_font():
        return Font(name="Calibri", bold=True, color=WHITE, size=11)

    def _title_font(size=14):
        return Font(name="Georgia", bold=True, color=DARK, size=size)

    # ── Title section ─────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"] = f"✦  {salon} — Booking Report"
    ws["A1"].font      = _title_font(18)
    ws["A1"].fill      = _fill(DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font      = Font(name="Georgia", bold=True, color=GOLD, size=16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Total: {summary['total']} bookings  |  Advance: {cur}{summary['advance_collected']:,}"
    ws["A2"].font      = Font(name="Calibri", color="666666", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8  # spacer

    # ── Column headers ────────────────────────────────────────────────
    headers = [
        "#", "Booking ID", "Client Name", "Phone",
        "Service", "Staff", "Date", "Slot",
        "Duration", "Total Price", "Advance", "Payment", "Status"
    ]
    col_widths = [4, 14, 18, 14, 18, 12, 12, 8, 10, 12, 12, 10, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = _header_font()
        cell.fill      = _fill(GOLD)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────────
    for i, b in enumerate(bookings, 1):
        row = 4 + i
        bg  = GRAY if i % 2 == 0 else WHITE
        conflict = b["conflict_flag"] == 1
        paid     = b["payment_status"] == "paid"

        row_data = [
            i,
            b["id"],
            b["client_name"],
            b["phone"] or "—",
            b["service"],
            b["staff"],
            b["date"],
            b["slot"],
            f"{b['duration']} min",
            f"{cur}{b['total_price']:,}",
            f"{cur}{b['advance_amount']:,}",
            "✅ Paid" if paid else "⏳ Pending",
            "⚠️ Conflict" if conflict else "✓ OK",
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = _fill(RED if conflict else bg)
            cell.border    = _border()
            cell.font      = Font(name="Calibri", size=10,
                                  color=WHITE if conflict else DARK)
            cell.alignment = Alignment(vertical="center",
                                        horizontal="center" if col in [1,7,8,9,12,13] else "left")

        ws.row_dimensions[row].height = 18

    # ── Summary section ───────────────────────────────────────────────
    sum_row = 4 + len(bookings) + 2

    summary_data = [
        ("Total Bookings",    summary["total"]),
        ("Total Revenue",     f"{cur}{summary['gross']:,}"),
        ("Advance Collected", f"{cur}{summary['advance_collected']:,}"),
        ("Paid Bookings",     summary["paid_count"]),
        ("Conflicts",         summary["conflicts"]),
    ]

    ws.cell(row=sum_row, column=1, value="SUMMARY").font = _title_font(12)
    ws.cell(row=sum_row, column=1).fill = _fill(DARK)
    ws.cell(row=sum_row, column=1).font = Font(name="Georgia", bold=True, color=GOLD, size=12)

    for idx, (label, val) in enumerate(summary_data):
        r = sum_row + 1 + idx
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        lc.font = Font(name="Calibri", bold=True, color=DARK, size=11)
        vc.font = Font(name="Calibri", color=DARK, size=11)
        lc.fill = vc.fill = _fill(LIGHT)
        lc.border = vc.border = _border()

    # ── Freeze header ──────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    wb.save(output_path)
    logger.info(f"[REPORTS] ✅ Report saved: {output_path}")
    return output_path
